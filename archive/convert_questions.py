# File Name: convert_questions.py
# Owner: Andrew
# Purpose: Converts CAPM questions from test_questions.xlsx to questions.json, appending non-destructively
# Version Control: v1.2
# Change Log: Updated: 05-Aug-2025, v1.2: Process all rows, check full questions.json for duplicates
import openpyxl
import json

def convert_questions(excel_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\test_questions.xlsx',
                     json_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\data\\questions.json'):
    try:
        # Load existing JSON data
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            last_id = max(item['id'] for item in json_data) if json_data else 301
            existing_questions = {item['question'] for item in json_data}
        except FileNotFoundError:
            json_data = []
            last_id = 301
            existing_questions = set()

        # Load Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Sheet1']
        new_questions = []
        
        # Process all rows starting from row 5
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            question_text = row[4]
            if question_text in existing_questions:
                print(f"Skipping duplicate question: {question_text}")
                continue
            new_questions.append({
                "id": last_id + 1,
                "type": row[1] if row[1] else "single",
                "domain": row[2] if row[2] else 4,
                "module": str(row[3]) if row[3] else "6",
                "question": question_text,
                "options": [f"a. {row[5]}", f"b. {row[6]}", f"c. {row[7]}", f"d. {row[8]}"],
                "answer": [int(x) for x in str(row[9]).split(',')] if ',' in str(row[9]) else int(row[9]) if row[9] else 0,
                "explanation": row[10] if row[10] else ""
            })
            last_id += 1
            existing_questions.add(question_text)

        # Append new questions to JSON
        json_data.extend(new_questions)
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=4)
        print(f"Added {len(new_questions)} questions to {json_file}")

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"Error converting questions: {str(e)}")

if __name__ == "__main__":
    convert_questions()