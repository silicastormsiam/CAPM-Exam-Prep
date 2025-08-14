# File Name: add_questions_to_excel.py
# Owner: Andrew
# Purpose: Appends CAPM questions from drop_the_format_questions.txt to test_questions.xlsx dynamically
# Version Control: v1.2
# Change Log: Updated: 05-Aug-2025, v1.2: Dynamic parsing, robust duplicate check, error handling
import openpyxl
import re

def parse_questions_from_txt(txt_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\data\\drop_the_format_questions.txt'):
    questions = []
    try:
        with open(txt_file, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            if not content:
                print(f"Error: {txt_file} is empty")
                return questions
            blocks = content.split('//')
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            lines = block.split('\n')
            question_data = {}
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if line.startswith('Question:'):
                    question_data['question'] = line.replace('Question:', '').strip()
                elif line.startswith('Option A:'):
                    question_data['option_a'] = line.replace('Option A:', '').strip()
                elif line.startswith('Option B:'):
                    question_data['option_b'] = line.replace('Option B:', '').strip()
                elif line.startswith('Option C:'):
                    question_data['option_c'] = line.replace('Option C:', '').strip()
                elif line.startswith('Option D:'):
                    question_data['option_d'] = line.replace('Option D:', '').strip()
                elif line.startswith('Answer:'):
                    answer = line.replace('Answer:', '').strip()
                    question_data['answer'] = answer
                    question_data['type'] = 'multiple' if ',' in answer else 'single'
                elif line.startswith('Explanation:'):
                    question_data['explanation'] = line.replace('Explanation:', '').strip()
            if len(question_data) >= 7:
                question_data['id'] = 0
                question_data['domain'] = 4
                question_data['module'] = '6'
                questions.append(question_data)
            else:
                print(f"Skipping incomplete question: {question_data.get('question', 'Unknown')}")
        print(f"Parsed {len(questions)} questions from {txt_file}")
        return questions
    except FileNotFoundError:
        print(f"Error: {txt_file} not found")
        return []
    except Exception as e:
        print(f"Error parsing {txt_file}: {str(e)}")
        return []

def append_questions(questions, excel_file='M:\\OneDrive\\Documents\\GitHub\\excel_to_json\\test_questions.xlsx'):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Sheet1']
        existing_questions = {ws[f'E{i}'].value for i in range(5, ws.max_row + 1) if ws[f'E{i}'].value}
        start_row = max(ws.max_row + 1, 5)
        new_count = 0
        for i, q in enumerate(questions, start=start_row):
            if q['question'] in existing_questions:
                print(f"Skipping duplicate question: {q['question']}")
                continue
            ws[f'A{i}'] = q['id']
            ws[f'B{i}'] = q['type']
            ws[f'C{i}'] = q['domain']
            ws[f'D{i}'] = q['module']
            ws[f'E{i}'] = q['question']
            ws[f'F{i}'] = q['option_a']
            ws[f'G{i}'] = q['option_b']
            ws[f'H{i}'] = q['option_c']
            ws[f'I{i}'] = q['option_d']
            ws[f'J{i}'] = q['answer']
            ws[f'K{i}'] = q['explanation']
            new_count += 1
            existing_questions.add(q['question'])
        wb.save(excel_file)
        print(f"Appended {new_count} CAPM questions to {excel_file}")
    except FileNotFoundError:
        print(f"Error: {excel_file} not found")
    except Exception as e:
        print(f"Error appending to {excel_file}: {str(e)}")

if __name__ == "__main__":
    questions = parse_questions_from_txt()
    if questions:
        append_questions(questions)
    else:
        print("No questions to append")